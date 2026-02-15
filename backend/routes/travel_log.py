"""
Travel Log Routes - Employee travel tracking and allowance management
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
from bson import ObjectId
from io import BytesIO
import os
import uuid
import shutil
import calendar
import base64
import re

# PDF/Excel
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# LLM for OCR
from dotenv import load_dotenv
load_dotenv()

router = APIRouter(prefix="/api/travel-log", tags=["Travel Log"])

# MongoDB connection
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME", "enerzia_erp")
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# Upload directory
UPLOADS_DIR = "/app/uploads/travel-photos"
os.makedirs(UPLOADS_DIR, exist_ok=True)


def serialize_doc(doc):
    """Convert MongoDB document to serializable dict"""
    if doc is None:
        return None
    doc = dict(doc)
    if "_id" in doc:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
    return doc


# ============= MODELS =============

class TravelTrip(BaseModel):
    from_location: str
    to_location: str
    vehicle_type: str  # two_wheeler, four_wheeler
    start_km: float
    end_km: float
    purpose: str  # Project Execution, Site Visit, Material Purchase, Client Meeting, Delivery, Other
    notes: Optional[str] = None


class TravelRates(BaseModel):
    two_wheeler_rate: float = 4.25
    four_wheeler_rate: float = 9.0


# ============= RATE SETTINGS =============

@router.get("/rates")
async def get_travel_rates():
    """Get current travel allowance rates"""
    rates = await db.travel_settings.find_one({"type": "rates"})
    if rates:
        return serialize_doc(rates)
    # Return defaults if not set
    return {
        "two_wheeler_rate": 4.25,
        "four_wheeler_rate": 9.0
    }


@router.put("/rates")
async def update_travel_rates(rates: TravelRates):
    """Update travel allowance rates (HR/Admin only)"""
    await db.travel_settings.update_one(
        {"type": "rates"},
        {"$set": {
            "type": "rates",
            "two_wheeler_rate": rates.two_wheeler_rate,
            "four_wheeler_rate": rates.four_wheeler_rate,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    return {"message": "Rates updated successfully", "rates": rates.dict()}


# ============= ODOMETER OCR =============

@router.post("/ocr/odometer")
async def extract_odometer_reading(
    photo: UploadFile = File(...)
):
    """Extract odometer reading from uploaded photo using AI vision"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        
        # Read the image file
        image_data = await photo.read()
        
        # Convert to base64
        image_base64 = base64.b64encode(image_data).decode('utf-8')
        
        # Get API key
        api_key = os.environ.get("EMERGENT_LLM_KEY")
        if not api_key:
            raise HTTPException(status_code=500, detail="OCR service not configured")
        
        # Initialize chat with Gemini Flash for fast OCR
        chat = LlmChat(
            api_key=api_key,
            session_id=f"ocr-{uuid.uuid4()}",
            system_message="You are an OCR assistant that extracts odometer readings from vehicle dashboard photos. Extract ONLY the numeric odometer reading (kilometers). Return ONLY the number, no units or text. If you cannot read the odometer clearly, return 'UNREADABLE'."
        ).with_model("gemini", "gemini-2.0-flash")
        
        # Create image content
        image_content = ImageContent(image_base64=image_base64)
        
        # Create message with image - use file_contents parameter
        user_message = UserMessage(
            text="Extract the odometer reading (km) from this vehicle dashboard photo. Return ONLY the numeric value, nothing else.",
            file_contents=[image_content]
        )
        
        # Get response
        response = await chat.send_message(user_message)
        
        # Parse the response to extract number
        response_text = response.strip() if response else ""
        
        if response_text == "UNREADABLE" or not response_text:
            return {
                "success": False,
                "error": "Could not read odometer value from image",
                "raw_response": response_text
            }
        
        # Extract numeric value from response
        # Remove any non-numeric characters except decimal point
        numeric_match = re.search(r'[\d,]+\.?\d*', response_text.replace(',', ''))
        if numeric_match:
            odometer_value = numeric_match.group().replace(',', '')
            return {
                "success": True,
                "odometer_reading": float(odometer_value),
                "raw_response": response_text
            }
        else:
            return {
                "success": False,
                "error": "Could not parse numeric value from response",
                "raw_response": response_text
            }
            
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"OCR library not available: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR failed: {str(e)}")


# ============= TRIP MANAGEMENT =============

@router.post("/trip")
async def create_trip(
    user_id: str = Form(...),
    user_name: str = Form(...),
    department: str = Form(None),
    from_location: str = Form(...),
    to_location: str = Form(...),
    vehicle_type: str = Form(...),
    start_km: float = Form(...),
    end_km: float = Form(...),
    purpose: str = Form(...),
    notes: str = Form(None),
    start_photo: UploadFile = File(None),
    end_photo: UploadFile = File(None)
):
    """Create a new travel trip entry"""
    
    # Calculate distance
    distance = end_km - start_km
    if distance < 0:
        raise HTTPException(status_code=400, detail="End KM must be greater than Start KM")
    
    # Get rates
    rates = await db.travel_settings.find_one({"type": "rates"})
    two_wheeler_rate = rates.get("two_wheeler_rate", 4.25) if rates else 4.25
    four_wheeler_rate = rates.get("four_wheeler_rate", 9.0) if rates else 9.0
    
    # Calculate allowance
    rate = two_wheeler_rate if vehicle_type == "two_wheeler" else four_wheeler_rate
    allowance = round(distance * rate, 2)
    
    # Handle photo uploads
    start_photo_path = None
    end_photo_path = None
    
    if start_photo:
        file_ext = start_photo.filename.split('.')[-1] if '.' in start_photo.filename else 'jpg'
        start_photo_filename = f"{uuid.uuid4()}_start.{file_ext}"
        start_photo_path = f"/travel-photos/{start_photo_filename}"
        with open(f"{UPLOADS_DIR}/{start_photo_filename}", "wb") as f:
            shutil.copyfileobj(start_photo.file, f)
    
    if end_photo:
        file_ext = end_photo.filename.split('.')[-1] if '.' in end_photo.filename else 'jpg'
        end_photo_filename = f"{uuid.uuid4()}_end.{file_ext}"
        end_photo_path = f"/travel-photos/{end_photo_filename}"
        with open(f"{UPLOADS_DIR}/{end_photo_filename}", "wb") as f:
            shutil.copyfileobj(end_photo.file, f)
    
    trip_doc = {
        "user_id": user_id,
        "user_name": user_name,
        "department": department,
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "time": datetime.now(timezone.utc).strftime("%H:%M"),
        "from_location": from_location,
        "to_location": to_location,
        "vehicle_type": vehicle_type,
        "start_km": start_km,
        "end_km": end_km,
        "distance": distance,
        "purpose": purpose,
        "notes": notes,
        "start_photo": start_photo_path,
        "end_photo": end_photo_path,
        "rate_applied": rate,
        "allowance": allowance,
        "status": "pending",  # pending, approved, rejected
        "approved_by": None,
        "approved_at": None,
        "rejection_reason": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.travel_logs.insert_one(trip_doc)
    trip_doc["id"] = str(result.inserted_id)
    if "_id" in trip_doc:
        del trip_doc["_id"]
    
    return {"message": "Trip logged successfully", "trip": trip_doc}


@router.get("/my-trips/{user_id}")
async def get_my_trips(user_id: str, date: Optional[str] = None, month: Optional[int] = None, year: Optional[int] = None):
    """Get trips for a specific user"""
    query = {"user_id": user_id}
    
    if date:
        query["date"] = date
    elif month and year:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query["date"] = {"$gte": start_date, "$lt": end_date}
    
    cursor = db.travel_logs.find(query).sort("created_at", -1)
    trips = []
    async for doc in cursor:
        trips.append(serialize_doc(doc))
    
    # Calculate summary
    total_distance = sum(t.get("distance", 0) for t in trips)
    total_allowance = sum(t.get("allowance", 0) for t in trips)
    approved_allowance = sum(t.get("allowance", 0) for t in trips if t.get("status") == "approved")
    pending_count = sum(1 for t in trips if t.get("status") == "pending")
    approved_count = sum(1 for t in trips if t.get("status") == "approved")
    rejected_count = sum(1 for t in trips if t.get("status") == "rejected")
    
    return {
        "trips": trips,
        "summary": {
            "total_trips": len(trips),
            "total_distance": round(total_distance, 2),
            "total_allowance": round(total_allowance, 2),
            "approved_allowance": round(approved_allowance, 2),
            "pending": pending_count,
            "approved": approved_count,
            "rejected": rejected_count
        }
    }


@router.get("/today/{user_id}")
async def get_today_trips(user_id: str):
    """Get today's trips for a user"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return await get_my_trips(user_id, date=today)


@router.delete("/trip/{trip_id}")
async def delete_trip(trip_id: str):
    """Delete a trip (only if pending)"""
    trip = await db.travel_logs.find_one({"_id": ObjectId(trip_id)})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Cannot delete approved/rejected trips")
    
    # Delete photos if exist
    if trip.get("start_photo"):
        photo_path = f"{UPLOADS_DIR}/{trip['start_photo'].split('/')[-1]}"
        if os.path.exists(photo_path):
            os.remove(photo_path)
    
    if trip.get("end_photo"):
        photo_path = f"{UPLOADS_DIR}/{trip['end_photo'].split('/')[-1]}"
        if os.path.exists(photo_path):
            os.remove(photo_path)
    
    await db.travel_logs.delete_one({"_id": ObjectId(trip_id)})
    return {"message": "Trip deleted successfully"}


# ============= HR MANAGEMENT =============

@router.get("/all-trips")
async def get_all_trips(
    status: Optional[str] = None,
    department: Optional[str] = None,
    user_id: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None
):
    """Get all trips for HR management"""
    query = {}
    
    if status:
        query["status"] = status
    
    if department:
        query["department"] = {"$regex": department, "$options": "i"}
    
    if user_id:
        query["user_id"] = user_id
    
    if month and year:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query["date"] = {"$gte": start_date, "$lt": end_date}
    
    cursor = db.travel_logs.find(query).sort("created_at", -1)
    trips = []
    async for doc in cursor:
        trips.append(serialize_doc(doc))
    
    # Calculate summary
    total_distance = sum(t.get("distance", 0) for t in trips)
    total_allowance = sum(t.get("allowance", 0) for t in trips)
    pending_count = sum(1 for t in trips if t.get("status") == "pending")
    approved_count = sum(1 for t in trips if t.get("status") == "approved")
    
    return {
        "trips": trips,
        "summary": {
            "total_trips": len(trips),
            "total_distance": round(total_distance, 2),
            "total_allowance": round(total_allowance, 2),
            "pending": pending_count,
            "approved": approved_count
        }
    }


@router.put("/trip/{trip_id}/approve")
async def approve_trip(trip_id: str, approved_by: str):
    """Approve a travel trip"""
    result = await db.travel_logs.update_one(
        {"_id": ObjectId(trip_id)},
        {"$set": {
            "status": "approved",
            "approved_by": approved_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Trip not found")
    return {"message": "Trip approved successfully"}


@router.put("/trip/{trip_id}/reject")
async def reject_trip(trip_id: str, rejected_by: str, reason: str = ""):
    """Reject a travel trip"""
    result = await db.travel_logs.update_one(
        {"_id": ObjectId(trip_id)},
        {"$set": {
            "status": "rejected",
            "approved_by": rejected_by,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "rejection_reason": reason
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Trip not found")
    return {"message": "Trip rejected"}


@router.put("/trip/{trip_id}/resubmit")
async def resubmit_trip(
    trip_id: str,
    start_location: str = Form(...),
    end_location: str = Form(...),
    start_km: float = Form(...),
    end_km: float = Form(...),
    purpose: str = Form(...),
    vehicle_type: str = Form(...),
    notes: str = Form(None),
    start_photo: UploadFile = File(None),
    end_photo: UploadFile = File(None)
):
    """Resubmit a rejected trip with updated details"""
    # Verify trip exists and is rejected
    trip = await db.travel_logs.find_one({"_id": ObjectId(trip_id)})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if trip.get("status") != "rejected":
        raise HTTPException(status_code=400, detail="Only rejected trips can be resubmitted")
    
    # Validate KM
    if end_km <= start_km:
        raise HTTPException(status_code=400, detail="End KM must be greater than Start KM")
    
    distance = round(end_km - start_km, 2)
    
    # Get rates
    rates_doc = await db.travel_settings.find_one({"type": "rates"})
    rates = rates_doc or {"two_wheeler_rate": 4.25, "four_wheeler_rate": 9.0}
    rate = rates.get("two_wheeler_rate", 4.25) if vehicle_type == "two_wheeler" else rates.get("four_wheeler_rate", 9.0)
    allowance = round(distance * rate, 2)
    
    # Handle photo uploads
    update_data = {
        "start_location": start_location,
        "end_location": end_location,
        "start_km": start_km,
        "end_km": end_km,
        "distance": distance,
        "purpose": purpose,
        "vehicle_type": vehicle_type,
        "rate": rate,
        "allowance": allowance,
        "notes": notes,
        "status": "pending",  # Reset to pending
        "approved_by": None,
        "approved_at": None,
        "rejection_reason": None,
        "resubmitted_at": datetime.now(timezone.utc).isoformat(),
        "resubmit_count": trip.get("resubmit_count", 0) + 1
    }
    
    # Handle photo uploads
    if start_photo:
        start_photo_path = f"/app/uploads/travel/{uuid.uuid4()}_{start_photo.filename}"
        os.makedirs(os.path.dirname(start_photo_path), exist_ok=True)
        with open(start_photo_path, "wb") as f:
            shutil.copyfileobj(start_photo.file, f)
        update_data["start_photo"] = start_photo_path
    
    if end_photo:
        end_photo_path = f"/app/uploads/travel/{uuid.uuid4()}_{end_photo.filename}"
        os.makedirs(os.path.dirname(end_photo_path), exist_ok=True)
        with open(end_photo_path, "wb") as f:
            shutil.copyfileobj(end_photo.file, f)
        update_data["end_photo"] = end_photo_path
    
    result = await db.travel_logs.update_one(
        {"_id": ObjectId(trip_id)},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Failed to update trip")
    
    return {
        "message": "Trip resubmitted for approval",
        "trip_id": trip_id,
        "distance": distance,
        "allowance": allowance
    }


@router.put("/bulk-approve")
async def bulk_approve_trips(trip_ids: List[str], approved_by: str):
    """Bulk approve multiple trips"""
    for trip_id in trip_ids:
        await db.travel_logs.update_one(
            {"_id": ObjectId(trip_id)},
            {"$set": {
                "status": "approved",
                "approved_by": approved_by,
                "approved_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    return {"message": f"{len(trip_ids)} trips approved"}


# ============= REPORTS =============

@router.get("/report/user/{user_id}")
async def get_user_travel_report(user_id: str, month: int, year: int):
    """Get travel report for a specific user"""
    data = await get_my_trips(user_id, month=month, year=year)
    
    # Get user info
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    
    return {
        "user": user,
        "trips": data["trips"],
        "summary": data["summary"],
        "month": month,
        "year": year
    }


@router.get("/report/download/pdf/{user_id}")
async def download_travel_report_pdf(user_id: str, month: int, year: int):
    """Download travel report as PDF"""
    data = await get_my_trips(user_id, month=month, year=year)
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    org_settings = await db.organization_settings.find_one({})
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Company Header
    company_name = org_settings.get('name', 'Workhub Enerzia') if org_settings else 'Workhub Enerzia'
    elements.append(Paragraph(company_name, ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER)))
    elements.append(Spacer(1, 10))
    
    month_name = calendar.month_name[month]
    elements.append(Paragraph(f"Travel Allowance Report - {month_name} {year}", ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=12, alignment=TA_CENTER)))
    elements.append(Spacer(1, 20))
    
    # Employee Details
    if user:
        emp_data = [
            ['Employee:', user.get('name', 'N/A'), 'Department:', user.get('department', 'N/A')],
        ]
        emp_table = Table(emp_data, colWidths=[70, 150, 70, 150])
        emp_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(emp_table)
        elements.append(Spacer(1, 15))
    
    # Trips Table
    table_data = [['#', 'Date', 'From → To', 'Vehicle', 'KM', 'Amount', 'Status']]
    
    for idx, trip in enumerate(data["trips"], 1):
        route = f"{trip.get('from_location', '')} → {trip.get('to_location', '')}"
        if len(route) > 25:
            route = route[:25] + "..."
        vehicle = "🏍️ 2W" if trip.get('vehicle_type') == 'two_wheeler' else "🚗 4W"
        status = trip.get('status', 'pending').capitalize()
        
        table_data.append([
            str(idx),
            trip.get('date', ''),
            route,
            vehicle,
            f"{trip.get('distance', 0):.1f}",
            f"₹{trip.get('allowance', 0):.2f}",
            status
        ])
    
    col_widths = [25, 60, 130, 40, 40, 60, 50]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Summary
    summary = data["summary"]
    summary_data = [
        ['Summary', '', ''],
        ['Total Trips:', str(summary['total_trips']), ''],
        ['Total Distance:', f"{summary['total_distance']:.2f} km", ''],
        ['Total Allowance:', f"₹{summary['total_allowance']:.2f}", ''],
        ['Approved Amount:', f"₹{summary['approved_allowance']:.2f}", ''],
    ]
    summary_table = Table(summary_data, colWidths=[120, 100, 100])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    
    # Footer
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')} | {company_name}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#94a3b8'), alignment=TA_CENTER)
    ))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"Travel_Report_{user.get('name', 'Employee')}_{month_name}_{year}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/report/download/excel")
async def download_all_travel_report_excel(month: int, year: int, department: Optional[str] = None):
    """Download all employees travel report as Excel"""
    data = await get_all_trips(month=month, year=year, department=department)
    org_settings = await db.organization_settings.find_one({})
    company_name = org_settings.get('name', 'Workhub Enerzia') if org_settings else 'Workhub Enerzia'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Travel Allowance Report"
    
    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    month_name = calendar.month_name[month]
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Travel Allowance Report - {month_name} {year}"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Table Headers
    headers = ['#', 'Employee', 'Department', 'Date', 'Route', 'Vehicle', 'Distance (km)', 'Allowance (₹)', 'Status']
    header_row = 5
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for idx, trip in enumerate(data["trips"], 1):
        row = header_row + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=trip.get('user_name', ''))
        ws.cell(row=row, column=3, value=trip.get('department', ''))
        ws.cell(row=row, column=4, value=trip.get('date', ''))
        ws.cell(row=row, column=5, value=f"{trip.get('from_location', '')} → {trip.get('to_location', '')}")
        ws.cell(row=row, column=6, value='2W' if trip.get('vehicle_type') == 'two_wheeler' else '4W')
        ws.cell(row=row, column=7, value=trip.get('distance', 0))
        ws.cell(row=row, column=8, value=trip.get('allowance', 0))
        ws.cell(row=row, column=9, value=trip.get('status', '').capitalize())
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"All_Travel_Report_{month_name}_{year}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
