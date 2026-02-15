"""
Sales Module API Routes
- Enquiries (Enq/Year/XXXX format)
- Quotations (linked to enquiries)
- Orders (linked to quotations)
- Sales Targets (FY-based, monthly/quarterly)
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
from bson import ObjectId
import uuid
import os
import io
from motor.motor_asyncio import AsyncIOMotorClient

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'enerzia_erp')]

router = APIRouter(prefix="/api/sales", tags=["Sales"])


# ============== MODELS ==============

class EnquiryItem(BaseModel):
    """Line item in an enquiry"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str
    quantity: Optional[float] = None
    unit: Optional[str] = None
    unit_price: Optional[float] = None
    total: Optional[float] = None


class EnquiryCreate(BaseModel):
    """Create a new enquiry"""
    date: str  # YYYY-MM-DD format
    target_date: Optional[str] = None  # YYYY-MM-DD format
    company_name: str
    customer_id: Optional[str] = None  # Link to existing customer
    location: Optional[str] = None
    description: str
    value: Optional[float] = None
    contact_person: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    remarks: Optional[str] = None
    category: Optional[str] = None  # PSS, AS, OSS, CS, DOM Lighting, Exports
    assigned_to: Optional[str] = None
    department: Optional[str] = None  # PROJECTS, SALES, ACCOUNTS, etc.
    priority: Optional[str] = None  # high, medium, low
    status: Optional[str] = None  # new, price_enquiry, site_visit_needed, etc.


class EnquiryUpdate(BaseModel):
    """Update an enquiry"""
    date: Optional[str] = None
    target_date: Optional[str] = None
    company_name: Optional[str] = None
    customer_id: Optional[str] = None
    location: Optional[str] = None
    description: Optional[str] = None
    value: Optional[float] = None
    contact_person: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    remarks: Optional[str] = None
    status: Optional[str] = None
    category: Optional[str] = None
    assigned_to: Optional[str] = None
    department: Optional[str] = None
    priority: Optional[str] = None


class QuotationItem(BaseModel):
    """Line item in a quotation"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sno: int
    description: str
    unit: str = "Nos"
    quantity: float = 1
    unit_price: float = 0
    total: float = 0


class QuotationCreate(BaseModel):
    """Create a new quotation"""
    enquiry_id: Optional[str] = None  # Link to enquiry
    customer_name: str
    customer_address: Optional[str] = None
    customer_gst: Optional[str] = None
    customer_contact: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    date: str  # DD/MM/YYYY
    valid_until: str  # DD/MM/YYYY
    subject: Optional[str] = None
    items: List[dict] = []
    subtotal: float = 0
    gst_percent: float = 18
    gst_amount: float = 0
    total_amount: float = 0
    terms_conditions: Optional[str] = None
    payment_terms: Optional[str] = None
    delivery_terms: Optional[str] = None
    notes: Optional[str] = None
    prepared_by: Optional[str] = None
    category: Optional[str] = None  # PSS, AS, OSS, CS


class QuotationUpdate(BaseModel):
    """Update a quotation"""
    customer_name: Optional[str] = None
    customer_address: Optional[str] = None
    customer_gst: Optional[str] = None
    customer_contact: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    date: Optional[str] = None
    valid_until: Optional[str] = None
    subject: Optional[str] = None
    items: Optional[List[dict]] = None
    subtotal: Optional[float] = None
    gst_percent: Optional[float] = None
    gst_amount: Optional[float] = None
    total_amount: Optional[float] = None
    terms_conditions: Optional[str] = None
    payment_terms: Optional[str] = None
    delivery_terms: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    prepared_by: Optional[str] = None
    category: Optional[str] = None


class OrderCreate(BaseModel):
    """Create a new sales order"""
    quotation_id: Optional[str] = None  # Link to quotation
    enquiry_id: Optional[str] = None  # Link to enquiry
    customer_name: str
    customer_address: Optional[str] = None
    customer_gst: Optional[str] = None
    customer_contact: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    date: str  # DD/MM/YYYY
    delivery_date: Optional[str] = None
    po_number: Optional[str] = None  # Customer PO number
    po_date: Optional[str] = None
    items: List[dict] = []
    subtotal: float = 0
    gst_percent: float = 18
    gst_amount: float = 0
    total_amount: float = 0
    payment_terms: Optional[str] = None
    delivery_terms: Optional[str] = None
    notes: Optional[str] = None
    category: Optional[str] = None


class OrderUpdate(BaseModel):
    """Update a sales order"""
    customer_name: Optional[str] = None
    customer_address: Optional[str] = None
    customer_gst: Optional[str] = None
    customer_contact: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    date: Optional[str] = None
    delivery_date: Optional[str] = None
    po_number: Optional[str] = None
    po_date: Optional[str] = None
    items: Optional[List[dict]] = None
    subtotal: Optional[float] = None
    gst_percent: Optional[float] = None
    gst_amount: Optional[float] = None
    total_amount: Optional[float] = None
    payment_terms: Optional[str] = None
    delivery_terms: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    payment_status: Optional[str] = None
    category: Optional[str] = None


class SalesTargetCreate(BaseModel):
    """Create/Update sales targets for a fiscal year"""
    fiscal_year: str  # e.g., "2024-25"
    category: str  # Domestic P&S, DOM Lighting, Exports
    monthly_targets: Dict[str, float] = {}  # {"apr": 7000000, "may": 7000000, ...}


# ============== HELPER FUNCTIONS ==============

async def get_next_enquiry_number():
    """Generate next enquiry number in format Enq/Year/XXXX"""
    current_year = datetime.now().year
    # Fiscal year format (Apr-Mar)
    current_month = datetime.now().month
    if current_month >= 4:  # Apr onwards
        fy_start = current_year
    else:  # Jan-Mar
        fy_start = current_year - 1
    
    fy_str = str(fy_start)  # Just use the year
    
    # Find the highest enquiry number for this fiscal year
    pattern = f"Enq/{fy_str}/"
    latest = await db.sales_enquiries.find_one(
        {"enquiry_no": {"$regex": f"^{pattern}"}},
        sort=[("enquiry_no", -1)]
    )
    
    if latest:
        try:
            last_num = int(latest["enquiry_no"].split("/")[-1])
            next_num = last_num + 1
        except (ValueError, IndexError):
            next_num = 1
    else:
        next_num = 1
    
    return f"Enq/{fy_str}/{str(next_num).zfill(4)}"


async def get_next_quotation_number():
    """Generate next quotation number"""
    current_year = datetime.now().year
    current_month = datetime.now().month
    if current_month >= 4:
        fy_start = current_year
    else:
        fy_start = current_year - 1
    
    fy_str = f"{fy_start % 100:02d}-{(fy_start + 1) % 100:02d}"
    
    pattern = f"QT-{fy_str}-"
    latest = await db.sales_quotations.find_one(
        {"quotation_no": {"$regex": f"^{pattern}"}},
        sort=[("quotation_no", -1)]
    )
    
    if latest:
        try:
            last_num = int(latest["quotation_no"].split("-")[-1])
            next_num = last_num + 1
        except (ValueError, IndexError):
            next_num = 1
    else:
        next_num = 1
    
    return f"QT-{fy_str}-{str(next_num).zfill(4)}"


async def get_next_order_number():
    """Generate next order number"""
    current_year = datetime.now().year
    current_month = datetime.now().month
    if current_month >= 4:
        fy_start = current_year
    else:
        fy_start = current_year - 1
    
    fy_str = f"{fy_start % 100:02d}-{(fy_start + 1) % 100:02d}"
    
    pattern = f"SO-{fy_str}-"
    latest = await db.sales_orders.find_one(
        {"order_no": {"$regex": f"^{pattern}"}},
        sort=[("order_no", -1)]
    )
    
    if latest:
        try:
            last_num = int(latest["order_no"].split("-")[-1])
            next_num = last_num + 1
        except (ValueError, IndexError):
            next_num = 1
    else:
        next_num = 1
    
    return f"SO-{fy_str}-{str(next_num).zfill(4)}"


# ============== ENQUIRY ENDPOINTS ==============

@router.get("/enquiries")
async def get_enquiries(
    status: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 100,
    skip: int = 0
):
    """Get all enquiries with optional filters"""
    query = {}
    
    if status:
        query["status"] = status
    if category:
        query["category"] = category
    if search:
        query["$or"] = [
            {"enquiry_no": {"$regex": search, "$options": "i"}},
            {"company_name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"contact_person": {"$regex": search, "$options": "i"}}
        ]
    
    cursor = db.sales_enquiries.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
    enquiries = await cursor.to_list(length=limit)
    total = await db.sales_enquiries.count_documents(query)
    
    return {"enquiries": enquiries, "total": total}


@router.get("/enquiries/stats")
async def get_enquiry_stats():
    """Get enquiry statistics"""
    total = await db.sales_enquiries.count_documents({})
    new_count = await db.sales_enquiries.count_documents({"status": "new"})
    quoted = await db.sales_enquiries.count_documents({"status": "quoted"})
    accepted = await db.sales_enquiries.count_documents({"status": "accepted"})
    declined = await db.sales_enquiries.count_documents({"status": "declined"})
    
    # This month's enquiries
    now = datetime.now()
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    this_month = await db.sales_enquiries.count_documents({
        "created_at": {"$gte": month_start}
    })
    
    # Total value of new enquiries
    pipeline = [
        {"$match": {"status": {"$in": ["new", "quoted", "site_visited"]}}},
        {"$group": {"_id": None, "total_value": {"$sum": {"$ifNull": ["$value", 0]}}}}
    ]
    result = await db.sales_enquiries.aggregate(pipeline).to_list(1)
    pipeline_value = result[0]["total_value"] if result else 0
    
    return {
        "total": total,
        "new": new_count,
        "quoted": quoted,
        "accepted": accepted,
        "declined": declined,
        "this_month": this_month,
        "pipeline_value": pipeline_value
    }


@router.get("/enquiries/{enquiry_id}")
async def get_enquiry(enquiry_id: str):
    """Get a single enquiry by ID"""
    enquiry = await db.sales_enquiries.find_one({"id": enquiry_id}, {"_id": 0})
    if not enquiry:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    return enquiry


@router.post("/enquiries")
async def create_enquiry(data: EnquiryCreate):
    """Create a new enquiry"""
    enquiry_no = await get_next_enquiry_number()
    
    enquiry = {
        "id": str(uuid.uuid4()),
        "enquiry_no": enquiry_no,
        "date": data.date,
        "target_date": data.target_date,
        "company_name": data.company_name,
        "customer_id": data.customer_id,
        "location": data.location,
        "description": data.description,
        "value": data.value,
        "contact_person": data.contact_person,
        "contact_phone": data.contact_phone,
        "contact_email": data.contact_email,
        "remarks": data.remarks,
        "category": data.category,
        "assigned_to": data.assigned_to,
        "department": data.department,
        "priority": data.priority,
        "status": data.status or "new",  # new, price_enquiry, site_visit_needed, etc.
        "quotation_id": None,
        "order_id": None,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.sales_enquiries.insert_one(enquiry)
    enquiry.pop("_id", None)
    
    return {"message": "Enquiry created successfully", "enquiry": enquiry}


@router.put("/enquiries/{enquiry_id}")
async def update_enquiry(enquiry_id: str, data: EnquiryUpdate):
    """Update an enquiry"""
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.sales_enquiries.update_one(
        {"id": enquiry_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    
    enquiry = await db.sales_enquiries.find_one({"id": enquiry_id}, {"_id": 0})
    return {"message": "Enquiry updated successfully", "enquiry": enquiry}


@router.delete("/enquiries/{enquiry_id}")
async def delete_enquiry(enquiry_id: str):
    """Delete an enquiry"""
    result = await db.sales_enquiries.delete_one({"id": enquiry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    return {"message": "Enquiry deleted successfully"}


@router.get("/enquiries/export/pdf")
async def export_enquiries_pdf(
    status: Optional[str] = None,
    search: Optional[str] = None
):
    """Export enquiries to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    # Build query
    query = {}
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"enquiry_no": {"$regex": search, "$options": "i"}},
            {"company_name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    cursor = db.sales_enquiries.find(query, {"_id": 0}).sort("created_at", -1).limit(500)
    enquiries = await cursor.to_list(length=500)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1)
    elements.append(Paragraph("Enquiries Report", title_style))
    elements.append(Spacer(1, 20))
    
    # Date
    date_style = ParagraphStyle('Date', parent=styles['Normal'], fontSize=10, alignment=1)
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style))
    elements.append(Spacer(1, 20))
    
    # Table data
    headers = ['Enquiry No', 'Date', 'Company', 'Description', 'Value (₹)', 'Priority', 'Status', 'Assigned To']
    data = [headers]
    
    for enq in enquiries:
        value = f"₹{enq.get('value', 0):,.0f}" if enq.get('value') else '-'
        desc = enq.get('description', '')[:40] + '...' if len(enq.get('description', '')) > 40 else enq.get('description', '')
        priority = enq.get('priority', '-').title() if enq.get('priority') else '-'
        data.append([
            enq.get('enquiry_no', ''),
            enq.get('date', ''),
            enq.get('company_name', ''),
            desc,
            value,
            priority,
            enq.get('status', '').replace('_', ' ').title(),
            enq.get('assigned_to', '-')
        ])
    
    # Create table
    table = Table(data, colWidths=[1.1*inch, 0.8*inch, 1.6*inch, 2.2*inch, 0.9*inch, 0.7*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=enquiries_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@router.get("/enquiries/export/excel")
async def export_enquiries_excel(
    status: Optional[str] = None,
    search: Optional[str] = None
):
    """Export enquiries to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Build query
    query = {}
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"enquiry_no": {"$regex": search, "$options": "i"}},
            {"company_name": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    cursor = db.sales_enquiries.find(query, {"_id": 0}).sort("created_at", -1).limit(1000)
    enquiries = await cursor.to_list(length=1000)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Enquiries"
    
    # Headers - Complete list of fields
    headers = ['Enquiry No', 'Date', 'Target Date', 'Company', 'Location', 'Description', 
               'Value (₹)', 'Contact Person', 'Phone', 'Email', 'Priority', 'Status', 
               'Category', 'Department', 'Assigned To', 'Remarks']
    
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Data rows
    for row, enq in enumerate(enquiries, 2):
        ws.cell(row=row, column=1, value=enq.get('enquiry_no', '')).border = thin_border
        ws.cell(row=row, column=2, value=enq.get('date', '')).border = thin_border
        ws.cell(row=row, column=3, value=enq.get('target_date', '')).border = thin_border
        ws.cell(row=row, column=4, value=enq.get('company_name', '')).border = thin_border
        ws.cell(row=row, column=5, value=enq.get('location', '')).border = thin_border
        ws.cell(row=row, column=6, value=enq.get('description', '')).border = thin_border
        ws.cell(row=row, column=7, value=enq.get('value', 0) or '').border = thin_border
        ws.cell(row=row, column=8, value=enq.get('contact_person', '')).border = thin_border
        ws.cell(row=row, column=9, value=enq.get('contact_phone', '')).border = thin_border
        ws.cell(row=row, column=10, value=enq.get('contact_email', '')).border = thin_border
        ws.cell(row=row, column=11, value=(enq.get('priority', '') or '').title()).border = thin_border
        ws.cell(row=row, column=12, value=enq.get('status', '').replace('_', ' ').title()).border = thin_border
        ws.cell(row=row, column=13, value=enq.get('category', '')).border = thin_border
        ws.cell(row=row, column=14, value=enq.get('department', '')).border = thin_border
        ws.cell(row=row, column=15, value=enq.get('assigned_to', '')).border = thin_border
        ws.cell(row=row, column=16, value=enq.get('remarks', '')).border = thin_border
    
    # Adjust column widths
    column_widths = [15, 12, 12, 25, 20, 40, 12, 18, 15, 25, 10, 15, 12, 15, 18, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=enquiries_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )


@router.get("/enquiries/bulk/template")
async def download_bulk_template():
    """Download Excel template for bulk enquiry upload"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Enquiries Template"
    
    # Headers - Complete list matching the form fields
    headers = [
        'Date (YYYY-MM-DD) *', 'Target Date (YYYY-MM-DD)', 'Company Name *', 'Location', 
        'Description *', 'Estimated Value', 'Contact Person', 'Contact Phone', 
        'Contact Email', 'Priority', 'Status', 'Category', 'Department', 'Assigned To', 'Remarks'
    ]
    
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add sample row
    sample_data = [
        datetime.now().strftime('%Y-%m-%d'),
        (datetime.now()).strftime('%Y-%m-%d'),
        'Sample Company Ltd',
        'Chennai, Tamil Nadu',
        'Enquiry for electrical services',
        '50000',
        'John Doe',
        '9876543210',
        'john@example.com',
        'high',
        'new',
        'PSS',
        'PROJECTS',
        'Team Member Name',
        'Sample remarks'
    ]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.font = Font(italic=True, color='666666')
    
    # Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        ["BULK UPLOAD INSTRUCTIONS"],
        [""],
        ["Required Fields (marked with *):", "Date, Company Name, Description"],
        [""],
        ["Date Format:", "YYYY-MM-DD (e.g., 2026-02-14)"],
        [""],
        ["Priority Options:", "high, medium, low"],
        [""],
        ["Status Options:", "new, price_enquiry, site_visit_needed, site_visited, under_progress, quoted, negotiation, accepted, declined, invoiced"],
        [""],
        ["Category Options:", "PSS, AS, OSS, CS, DOM_LIGHTING, EXPORTS"],
        [""],
        ["Department Options:", "PROJECTS, SALES, ACCOUNTS, PURCHASE, EXPORTS, FINANCE, HR, OPERATIONS"],
        [""],
        ["Notes:"],
        ["- Delete the sample row before uploading"],
        ["- Maximum 500 rows per upload"],
        ["- Duplicate company names are allowed"],
    ]
    
    for row, data in enumerate(instructions, 1):
        for col, value in enumerate(data, 1):
            ws_inst.cell(row=row, column=col, value=value)
    
    ws_inst.column_dimensions['A'].width = 30
    ws_inst.column_dimensions['B'].width = 50
    
    # Adjust column widths for main sheet
    column_widths = [18, 18, 25, 25, 40, 15, 20, 15, 25, 15, 15, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enquiries_bulk_upload_template.xlsx"}
    )


@router.post("/enquiries/bulk/upload")
async def bulk_upload_enquiries(file: UploadFile = File(...)):
    """Upload Excel file with multiple enquiries"""
    from openpyxl import load_workbook
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        imported = 0
        errors = []
        
        # Skip header row - Column order matches template:
        # Date, Target Date, Company Name, Location, Description, Value, 
        # Contact Person, Phone, Email, Priority, Status, Category, Department, Assigned To, Remarks
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue
            
            # Extract values (matching new template column order)
            date_val = row[0] if len(row) > 0 else None
            target_date_val = row[1] if len(row) > 1 else None
            company_name = row[2] if len(row) > 2 else None
            location = row[3] if len(row) > 3 else None
            description = row[4] if len(row) > 4 else None
            value = row[5] if len(row) > 5 else None
            contact_person = row[6] if len(row) > 6 else None
            contact_phone = row[7] if len(row) > 7 else None
            contact_email = row[8] if len(row) > 8 else None
            priority = row[9] if len(row) > 9 else None
            status = row[10] if len(row) > 10 else None
            category = row[11] if len(row) > 11 else None
            department = row[12] if len(row) > 12 else None
            assigned_to = row[13] if len(row) > 13 else None
            remarks = row[14] if len(row) > 14 else None
            
            # Validate required fields
            if not company_name or not description:
                errors.append(f"Row {row_num}: Company Name and Description are required")
                continue
            
            # Format date
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            elif date_val:
                date_str = str(date_val)
            else:
                date_str = datetime.now().strftime('%Y-%m-%d')
            
            if isinstance(target_date_val, datetime):
                target_date_str = target_date_val.strftime('%Y-%m-%d')
            else:
                target_date_str = str(target_date_val) if target_date_val else None
            
            # Parse value
            try:
                value_float = float(value) if value else None
            except (ValueError, TypeError):
                value_float = None
            
            # Validate and set status
            valid_statuses = ['new', 'price_enquiry', 'site_visit_needed', 'site_visited', 
                           'under_progress', 'quoted', 'negotiation', 'accepted', 'declined', 'invoiced']
            status_str = str(status).lower() if status else 'new'
            if status_str not in valid_statuses:
                status_str = 'new'
            
            # Validate priority
            valid_priorities = ['high', 'medium', 'low']
            priority_str = str(priority).lower() if priority else None
            if priority_str and priority_str not in valid_priorities:
                priority_str = None
            
            # Generate enquiry number
            enquiry_no = await get_next_enquiry_number()
            
            enquiry = {
                "id": str(uuid.uuid4()),
                "enquiry_no": enquiry_no,
                "date": date_str,
                "target_date": target_date_str,
                "company_name": str(company_name),
                "customer_id": None,
                "location": str(location) if location else None,
                "description": str(description),
                "value": value_float,
                "contact_person": str(contact_person) if contact_person else None,
                "contact_phone": str(contact_phone) if contact_phone else None,
                "contact_email": str(contact_email) if contact_email else None,
                "remarks": str(remarks) if remarks else None,
                "priority": priority_str,
                "status": status_str,
                "category": str(category) if category else None,
                "department": str(department) if department else None,
                "assigned_to": str(assigned_to) if assigned_to else None,
                "quotation_id": None,
                "order_id": None,
                "created_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }
            
            await db.sales_enquiries.insert_one(enquiry)
            imported += 1
        
        return {
            "message": f"Successfully imported {imported} enquiries",
            "imported": imported,
            "errors": errors[:10]  # Return first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


# ============== QUOTATION ENDPOINTS ==============

@router.get("/quotations")
async def get_quotations(
    status: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 100,
    skip: int = 0
):
    """Get all quotations"""
    query = {}
    
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"quotation_no": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}}
        ]
    
    cursor = db.sales_quotations.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
    quotations = await cursor.to_list(length=limit)
    total = await db.sales_quotations.count_documents(query)
    
    return {"quotations": quotations, "total": total}


@router.get("/quotations/stats")
async def get_quotation_stats():
    """Get quotation statistics"""
    total = await db.sales_quotations.count_documents({})
    draft = await db.sales_quotations.count_documents({"status": "draft"})
    sent = await db.sales_quotations.count_documents({"status": "sent"})
    accepted = await db.sales_quotations.count_documents({"status": "accepted"})
    rejected = await db.sales_quotations.count_documents({"status": "rejected"})
    
    # Total value
    pipeline = [
        {"$group": {"_id": None, "total_value": {"$sum": {"$ifNull": ["$total_amount", 0]}}}}
    ]
    result = await db.sales_quotations.aggregate(pipeline).to_list(1)
    total_value = result[0]["total_value"] if result else 0
    
    return {
        "total": total,
        "draft": draft,
        "sent": sent,
        "accepted": accepted,
        "rejected": rejected,
        "total_value": total_value
    }


@router.get("/quotations/{quotation_id}")
async def get_quotation(quotation_id: str):
    """Get a single quotation"""
    quotation = await db.sales_quotations.find_one({"id": quotation_id}, {"_id": 0})
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    return quotation


@router.post("/quotations")
async def create_quotation(data: QuotationCreate):
    """Create a new quotation"""
    quotation_no = await get_next_quotation_number()
    
    quotation = {
        "id": str(uuid.uuid4()),
        "quotation_no": quotation_no,
        "enquiry_id": data.enquiry_id,
        "customer_name": data.customer_name,
        "customer_address": data.customer_address,
        "customer_gst": data.customer_gst,
        "customer_contact": data.customer_contact,
        "customer_phone": data.customer_phone,
        "customer_email": data.customer_email,
        "date": data.date,
        "valid_until": data.valid_until,
        "subject": data.subject,
        "items": data.items,
        "subtotal": data.subtotal,
        "gst_percent": data.gst_percent,
        "gst_amount": data.gst_amount,
        "total_amount": data.total_amount,
        "terms_conditions": data.terms_conditions,
        "payment_terms": data.payment_terms,
        "delivery_terms": data.delivery_terms,
        "notes": data.notes,
        "prepared_by": data.prepared_by,
        "category": data.category,
        "status": "draft",  # draft, sent, accepted, rejected, expired
        "order_id": None,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.sales_quotations.insert_one(quotation)
    
    # Update enquiry if linked
    if data.enquiry_id:
        await db.sales_enquiries.update_one(
            {"id": data.enquiry_id},
            {"$set": {"quotation_id": quotation["id"], "status": "quoted", "updated_at": datetime.now(timezone.utc)}}
        )
    
    quotation.pop("_id", None)
    return {"message": "Quotation created successfully", "quotation": quotation}


@router.put("/quotations/{quotation_id}")
async def update_quotation(quotation_id: str, data: QuotationUpdate):
    """Update a quotation"""
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.sales_quotations.update_one(
        {"id": quotation_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    quotation = await db.sales_quotations.find_one({"id": quotation_id}, {"_id": 0})
    return {"message": "Quotation updated successfully", "quotation": quotation}


@router.delete("/quotations/{quotation_id}")
async def delete_quotation(quotation_id: str):
    """Delete a quotation"""
    quotation = await db.sales_quotations.find_one({"id": quotation_id})
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Unlink from enquiry if linked
    if quotation.get("enquiry_id"):
        await db.sales_enquiries.update_one(
            {"id": quotation["enquiry_id"]},
            {"$set": {"quotation_id": None, "status": "new", "updated_at": datetime.now(timezone.utc)}}
        )
    
    await db.sales_quotations.delete_one({"id": quotation_id})
    return {"message": "Quotation deleted successfully"}


@router.post("/quotations/{quotation_id}/convert-to-order")
async def convert_quotation_to_order(quotation_id: str):
    """Convert an accepted quotation to a sales order"""
    quotation = await db.sales_quotations.find_one({"id": quotation_id}, {"_id": 0})
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    if quotation.get("order_id"):
        raise HTTPException(status_code=400, detail="Quotation already converted to order")
    
    order_no = await get_next_order_number()
    
    order = {
        "id": str(uuid.uuid4()),
        "order_no": order_no,
        "quotation_id": quotation_id,
        "enquiry_id": quotation.get("enquiry_id"),
        "customer_name": quotation["customer_name"],
        "customer_address": quotation.get("customer_address"),
        "customer_gst": quotation.get("customer_gst"),
        "customer_contact": quotation.get("customer_contact"),
        "customer_phone": quotation.get("customer_phone"),
        "customer_email": quotation.get("customer_email"),
        "date": datetime.now().strftime("%d/%m/%Y"),
        "delivery_date": None,
        "po_number": None,
        "po_date": None,
        "items": quotation.get("items", []),
        "subtotal": quotation.get("subtotal", 0),
        "gst_percent": quotation.get("gst_percent", 18),
        "gst_amount": quotation.get("gst_amount", 0),
        "total_amount": quotation.get("total_amount", 0),
        "payment_terms": quotation.get("payment_terms"),
        "delivery_terms": quotation.get("delivery_terms"),
        "notes": quotation.get("notes"),
        "category": quotation.get("category"),
        "status": "pending",  # pending, confirmed, processing, shipped, delivered, cancelled
        "payment_status": "unpaid",  # unpaid, partial, paid
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.sales_orders.insert_one(order)
    
    # Update quotation
    await db.sales_quotations.update_one(
        {"id": quotation_id},
        {"$set": {"order_id": order["id"], "status": "accepted", "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Update enquiry if linked
    if quotation.get("enquiry_id"):
        await db.sales_enquiries.update_one(
            {"id": quotation["enquiry_id"]},
            {"$set": {"order_id": order["id"], "status": "accepted", "updated_at": datetime.now(timezone.utc)}}
        )
    
    order.pop("_id", None)
    return {"message": "Order created from quotation", "order": order}


# ============== ORDER ENDPOINTS ==============

@router.get("/orders")
async def get_orders(
    status: Optional[str] = None,
    payment_status: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 100,
    skip: int = 0
):
    """Get all sales orders"""
    query = {}
    
    if status:
        query["status"] = status
    if payment_status:
        query["payment_status"] = payment_status
    if search:
        query["$or"] = [
            {"order_no": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"po_number": {"$regex": search, "$options": "i"}}
        ]
    
    cursor = db.sales_orders.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit)
    orders = await cursor.to_list(length=limit)
    total = await db.sales_orders.count_documents(query)
    
    return {"orders": orders, "total": total}


@router.get("/orders/stats")
async def get_order_stats():
    """Get order statistics"""
    total = await db.sales_orders.count_documents({})
    pending = await db.sales_orders.count_documents({"status": "pending"})
    confirmed = await db.sales_orders.count_documents({"status": "confirmed"})
    processing = await db.sales_orders.count_documents({"status": "processing"})
    delivered = await db.sales_orders.count_documents({"status": "delivered"})
    
    # Total value
    pipeline = [
        {"$match": {"status": {"$ne": "cancelled"}}},
        {"$group": {"_id": None, "total_value": {"$sum": {"$ifNull": ["$total_amount", 0]}}}}
    ]
    result = await db.sales_orders.aggregate(pipeline).to_list(1)
    total_value = result[0]["total_value"] if result else 0
    
    # Paid amount
    pipeline_paid = [
        {"$match": {"payment_status": "paid"}},
        {"$group": {"_id": None, "paid_value": {"$sum": {"$ifNull": ["$total_amount", 0]}}}}
    ]
    result_paid = await db.sales_orders.aggregate(pipeline_paid).to_list(1)
    paid_value = result_paid[0]["paid_value"] if result_paid else 0
    
    return {
        "total": total,
        "pending": pending,
        "confirmed": confirmed,
        "processing": processing,
        "delivered": delivered,
        "total_value": total_value,
        "paid_value": paid_value
    }


@router.get("/orders/{order_id}")
async def get_order(order_id: str):
    """Get a single order"""
    order = await db.sales_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.post("/orders")
async def create_order(data: OrderCreate):
    """Create a new sales order"""
    order_no = await get_next_order_number()
    
    order = {
        "id": str(uuid.uuid4()),
        "order_no": order_no,
        "quotation_id": data.quotation_id,
        "enquiry_id": data.enquiry_id,
        "customer_name": data.customer_name,
        "customer_address": data.customer_address,
        "customer_gst": data.customer_gst,
        "customer_contact": data.customer_contact,
        "customer_phone": data.customer_phone,
        "customer_email": data.customer_email,
        "date": data.date,
        "delivery_date": data.delivery_date,
        "po_number": data.po_number,
        "po_date": data.po_date,
        "items": data.items,
        "subtotal": data.subtotal,
        "gst_percent": data.gst_percent,
        "gst_amount": data.gst_amount,
        "total_amount": data.total_amount,
        "payment_terms": data.payment_terms,
        "delivery_terms": data.delivery_terms,
        "notes": data.notes,
        "category": data.category,
        "status": "pending",
        "payment_status": "unpaid",
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.sales_orders.insert_one(order)
    order.pop("_id", None)
    
    return {"message": "Order created successfully", "order": order}


@router.put("/orders/{order_id}")
async def update_order(order_id: str, data: OrderUpdate):
    """Update a sales order"""
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    result = await db.sales_orders.update_one(
        {"id": order_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order = await db.sales_orders.find_one({"id": order_id}, {"_id": 0})
    return {"message": "Order updated successfully", "order": order}


@router.delete("/orders/{order_id}")
async def delete_order(order_id: str):
    """Delete a sales order"""
    order = await db.sales_orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Unlink from quotation if linked
    if order.get("quotation_id"):
        await db.sales_quotations.update_one(
            {"id": order["quotation_id"]},
            {"$set": {"order_id": None, "updated_at": datetime.now(timezone.utc)}}
        )
    
    # Unlink from enquiry if linked
    if order.get("enquiry_id"):
        await db.sales_enquiries.update_one(
            {"id": order["enquiry_id"]},
            {"$set": {"order_id": None, "updated_at": datetime.now(timezone.utc)}}
        )
    
    await db.sales_orders.delete_one({"id": order_id})
    return {"message": "Order deleted successfully"}


# ============== SALES TARGETS ENDPOINTS ==============

@router.get("/targets")
async def get_sales_targets(fiscal_year: Optional[str] = None):
    """Get sales targets"""
    query = {}
    if fiscal_year:
        query["fiscal_year"] = fiscal_year
    
    cursor = db.sales_targets.find(query, {"_id": 0}).sort("fiscal_year", -1)
    targets = await cursor.to_list(length=100)
    
    return {"targets": targets}


@router.get("/targets/{fiscal_year}")
async def get_fiscal_year_targets(fiscal_year: str):
    """Get targets for a specific fiscal year"""
    targets = await db.sales_targets.find({"fiscal_year": fiscal_year}, {"_id": 0}).to_list(length=100)
    return {"targets": targets, "fiscal_year": fiscal_year}


@router.post("/targets")
async def create_or_update_target(data: SalesTargetCreate):
    """Create or update sales target for a category in a fiscal year"""
    existing = await db.sales_targets.find_one({
        "fiscal_year": data.fiscal_year,
        "category": data.category
    })
    
    target_data = {
        "fiscal_year": data.fiscal_year,
        "category": data.category,
        "monthly_targets": data.monthly_targets,
        "updated_at": datetime.now(timezone.utc)
    }
    
    if existing:
        await db.sales_targets.update_one(
            {"fiscal_year": data.fiscal_year, "category": data.category},
            {"$set": target_data}
        )
        message = "Target updated successfully"
    else:
        target_data["id"] = str(uuid.uuid4())
        target_data["created_at"] = datetime.now(timezone.utc)
        await db.sales_targets.insert_one(target_data)
        message = "Target created successfully"
    
    target = await db.sales_targets.find_one(
        {"fiscal_year": data.fiscal_year, "category": data.category},
        {"_id": 0}
    )
    
    return {"message": message, "target": target}


@router.get("/targets/achievements/{fiscal_year}")
async def get_target_achievements(fiscal_year: str):
    """Get target vs achievement for a fiscal year"""
    # Get targets
    targets = await db.sales_targets.find({"fiscal_year": fiscal_year}, {"_id": 0}).to_list(length=100)
    
    # Calculate achievements from orders
    # Parse fiscal year to get date range
    try:
        fy_parts = fiscal_year.split("-")
        fy_start_year = int(fy_parts[0])
        if fy_start_year < 100:
            fy_start_year += 2000
        start_date = datetime(fy_start_year, 4, 1, tzinfo=timezone.utc)
        end_date = datetime(fy_start_year + 1, 3, 31, 23, 59, 59, tzinfo=timezone.utc)
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="Invalid fiscal year format. Use YY-YY or YYYY-YY")
    
    # Get orders in this fiscal year grouped by month and category
    pipeline = [
        {"$match": {
            "created_at": {"$gte": start_date, "$lte": end_date},
            "status": {"$ne": "cancelled"}
        }},
        {"$group": {
            "_id": {
                "month": {"$month": "$created_at"},
                "category": "$category"
            },
            "total": {"$sum": "$total_amount"}
        }}
    ]
    
    achievements_raw = await db.sales_orders.aggregate(pipeline).to_list(length=100)
    
    # Format achievements
    month_names = ["", "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    achievements = {}
    for item in achievements_raw:
        cat = item["_id"]["category"] or "Uncategorized"
        month = month_names[item["_id"]["month"]]
        if cat not in achievements:
            achievements[cat] = {}
        achievements[cat][month] = item["total"]
    
    return {
        "fiscal_year": fiscal_year,
        "targets": targets,
        "achievements": achievements
    }


# ============== DASHBOARD STATS ==============

@router.get("/dashboard/stats")
async def get_dashboard_stats():
    """Get overall sales dashboard statistics"""
    # Enquiry stats
    total_enquiries = await db.sales_enquiries.count_documents({})
    new_enquiries = await db.sales_enquiries.count_documents({"status": "new"})
    
    # Quotation stats
    total_quotations = await db.sales_quotations.count_documents({})
    active_quotations = await db.sales_quotations.count_documents({"status": {"$in": ["draft", "sent"]}})
    
    # Order stats
    total_orders = await db.sales_orders.count_documents({})
    pending_orders = await db.sales_orders.count_documents({"status": {"$in": ["pending", "confirmed", "processing"]}})
    
    # Revenue this month
    now = datetime.now()
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    pipeline = [
        {"$match": {
            "created_at": {"$gte": month_start},
            "status": {"$ne": "cancelled"}
        }},
        {"$group": {"_id": None, "revenue": {"$sum": "$total_amount"}}}
    ]
    result = await db.sales_orders.aggregate(pipeline).to_list(1)
    monthly_revenue = result[0]["revenue"] if result else 0
    
    # Conversion rate (enquiries to orders)
    total_with_orders = await db.sales_enquiries.count_documents({"order_id": {"$ne": None}})
    conversion_rate = (total_with_orders / total_enquiries * 100) if total_enquiries > 0 else 0
    
    return {
        "total_enquiries": total_enquiries,
        "new_enquiries": new_enquiries,
        "total_quotations": total_quotations,
        "active_quotations": active_quotations,
        "total_orders": total_orders,
        "pending_orders": pending_orders,
        "monthly_revenue": monthly_revenue,
        "conversion_rate": round(conversion_rate, 1)
    }
