"""
Attendance Reports - PDF and Excel generation for attendance data
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime
from io import BytesIO
import os
import calendar

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel Generation
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/attendance-reports", tags=["Attendance Reports"])

# MongoDB connection
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME", "enerzia_erp")
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]


def serialize_doc(doc):
    """Convert MongoDB document to serializable dict"""
    if doc is None:
        return None
    doc = dict(doc)
    if "_id" in doc:
        doc["id"] = str(doc["_id"])
        del doc["_id"]
    return doc


async def get_org_settings():
    """Get organization settings for report header"""
    settings = await db.organization_settings.find_one({})
    if settings:
        return serialize_doc(settings)
    return {
        "name": "Smarthub Enerzia",
        "address": "",
        "phone": "",
        "email": "",
        "logo_url": None
    }


async def get_attendance_data(user_id: str, month: int, year: int):
    """Fetch attendance data for a user"""
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    cursor = db.attendance.find({
        "user_id": user_id,
        "date": {"$gte": start_date, "$lt": end_date}
    }).sort("date", 1)
    
    records = []
    async for doc in cursor:
        records.append(serialize_doc(doc))
    
    # Get user details
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    
    return records, user


async def get_all_users_attendance(month: int, year: int, department: Optional[str] = None):
    """Fetch attendance data for all users or by department"""
    query = {}
    if department:
        query["department"] = {"$regex": department, "$options": "i"}
    
    users = await db.users.find(query, {"_id": 0, "password": 0}).to_list(1000)
    
    all_data = []
    for user in users:
        records, _ = await get_attendance_data(user.get("id"), month, year)
        
        present = sum(1 for r in records if r.get("status") == "present")
        absent = sum(1 for r in records if r.get("status") == "absent")
        half_days = sum(1 for r in records if r.get("status") == "half-day")
        on_leave = sum(1 for r in records if r.get("status") == "on-leave")
        
        all_data.append({
            "user": user,
            "records": records,
            "summary": {
                "present": present,
                "absent": absent,
                "half_days": half_days,
                "on_leave": on_leave,
                "total_days": len(records)
            }
        })
    
    return all_data


def generate_attendance_pdf(records, user, org_settings, month, year):
    """Generate PDF attendance report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#1e293b'),
        fontName='Helvetica-Bold'
    )
    
    # Company Header
    company_name = org_settings.get('name', 'Smarthub Enerzia')
    company_address = org_settings.get('address', '')
    company_phone = org_settings.get('phone', '')
    company_email = org_settings.get('email', '')
    
    elements.append(Paragraph(company_name, title_style))
    if company_address:
        elements.append(Paragraph(company_address, subtitle_style))
    
    contact_info = []
    if company_phone:
        contact_info.append(f"Phone: {company_phone}")
    if company_email:
        contact_info.append(f"Email: {company_email}")
    if contact_info:
        elements.append(Paragraph(" | ".join(contact_info), subtitle_style))
    
    elements.append(Spacer(1, 10))
    
    # Report Title
    month_name = calendar.month_name[month]
    elements.append(Paragraph(f"Attendance Report - {month_name} {year}", ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e293b'),
        alignment=TA_CENTER,
        spaceAfter=20
    )))
    
    # Employee Details
    if user:
        emp_name = user.get('name', 'N/A')
        emp_email = user.get('email', 'N/A')
        emp_dept = user.get('department', 'N/A')
        emp_id = user.get('id', 'N/A')[:8]
        
        emp_data = [
            ['Employee Name:', emp_name, 'Employee ID:', emp_id],
            ['Email:', emp_email, 'Department:', emp_dept],
        ]
        
        emp_table = Table(emp_data, colWidths=[80, 150, 80, 150])
        emp_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#374151')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(emp_table)
        elements.append(Spacer(1, 20))
    
    # Attendance Table
    table_data = [['#', 'Date', 'Day', 'Check In', 'Check Out', 'Status']]
    
    for idx, record in enumerate(records, 1):
        date_str = record.get('date', '')
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            day_name = date_obj.strftime('%A')
            formatted_date = date_obj.strftime('%d-%m-%Y')
        except (ValueError, TypeError):
            day_name = ''
            formatted_date = date_str
        
        check_in = record.get('check_in', '-')
        check_out = record.get('check_out', '-')
        status = record.get('status', 'N/A').capitalize()
        
        table_data.append([str(idx), formatted_date, day_name, check_in or '-', check_out or '-', status])
    
    # Create table with appropriate column widths
    col_widths = [30, 80, 70, 60, 60, 70]
    table = Table(table_data, colWidths=col_widths)
    
    # Table styling
    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (3, 1), (4, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])
    
    # Color-code status
    for idx, row in enumerate(table_data[1:], 1):
        status = row[5].lower()
        if status == 'present':
            table_style.add('TEXTCOLOR', (5, idx), (5, idx), colors.HexColor('#059669'))
        elif status == 'absent':
            table_style.add('TEXTCOLOR', (5, idx), (5, idx), colors.HexColor('#dc2626'))
        elif status == 'half-day':
            table_style.add('TEXTCOLOR', (5, idx), (5, idx), colors.HexColor('#d97706'))
        elif status == 'on-leave':
            table_style.add('TEXTCOLOR', (5, idx), (5, idx), colors.HexColor('#2563eb'))
    
    table.setStyle(table_style)
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Summary Section
    present = sum(1 for r in records if r.get("status") == "present")
    absent = sum(1 for r in records if r.get("status") == "absent")
    half_days = sum(1 for r in records if r.get("status") == "half-day")
    on_leave = sum(1 for r in records if r.get("status") == "on-leave")
    total = len(records)
    
    summary_data = [
        ['Summary', '', '', ''],
        ['Present Days:', str(present), 'Absent Days:', str(absent)],
        ['Half Days:', str(half_days), 'On Leave:', str(on_leave)],
        ['Total Working Days:', str(total), '', ''],
    ]
    
    summary_table = Table(summary_data, colWidths=[100, 60, 100, 60])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(summary_table)
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')} | {company_name}"
    elements.append(Paragraph(footer_text, ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#94a3b8'),
        alignment=TA_CENTER
    )))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_attendance_excel(records, user, org_settings, month, year):
    """Generate Excel attendance report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Styles
    header_font = Font(bold=True, size=14, color="1e293b")
    subheader_font = Font(size=10, color="64748b")
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    table_header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    present_font = Font(color="059669")
    absent_font = Font(color="dc2626")
    half_day_font = Font(color="d97706")
    leave_font = Font(color="2563eb")
    
    # Company Header
    company_name = org_settings.get('name', 'Smarthub Enerzia')
    ws.merge_cells('A1:F1')
    ws['A1'] = company_name
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    company_address = org_settings.get('address', '')
    if company_address:
        ws.merge_cells('A2:F2')
        ws['A2'] = company_address
        ws['A2'].font = subheader_font
        ws['A2'].alignment = Alignment(horizontal='center')
    
    # Report Title
    month_name = calendar.month_name[month]
    ws.merge_cells('A4:F4')
    ws['A4'] = f"Attendance Report - {month_name} {year}"
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    # Employee Details
    if user:
        ws['A6'] = "Employee Name:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = user.get('name', 'N/A')
        
        ws['D6'] = "Employee ID:"
        ws['D6'].font = Font(bold=True)
        ws['E6'] = user.get('id', 'N/A')[:8]
        
        ws['A7'] = "Email:"
        ws['A7'].font = Font(bold=True)
        ws['B7'] = user.get('email', 'N/A')
        
        ws['D7'] = "Department:"
        ws['D7'].font = Font(bold=True)
        ws['E7'] = user.get('department', 'N/A')
    
    # Table Headers
    headers = ['#', 'Date', 'Day', 'Check In', 'Check Out', 'Status']
    header_row = 9
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Table Data
    for idx, record in enumerate(records, 1):
        row = header_row + idx
        date_str = record.get('date', '')
        
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            day_name = date_obj.strftime('%A')
            formatted_date = date_obj.strftime('%d-%m-%Y')
        except (ValueError, TypeError):
            day_name = ''
            formatted_date = date_str
        
        check_in = record.get('check_in', '-') or '-'
        check_out = record.get('check_out', '-') or '-'
        status = record.get('status', 'N/A').capitalize()
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=formatted_date).border = thin_border
        ws.cell(row=row, column=3, value=day_name).border = thin_border
        ws.cell(row=row, column=4, value=check_in).border = thin_border
        ws.cell(row=row, column=5, value=check_out).border = thin_border
        
        status_cell = ws.cell(row=row, column=6, value=status)
        status_cell.border = thin_border
        
        # Color-code status
        if status.lower() == 'present':
            status_cell.font = present_font
        elif status.lower() == 'absent':
            status_cell.font = absent_font
        elif status.lower() == 'half-day':
            status_cell.font = half_day_font
        elif status.lower() == 'on-leave':
            status_cell.font = leave_font
        
        # Center alignment for certain columns
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
    
    # Summary Section
    summary_row = header_row + len(records) + 2
    present = sum(1 for r in records if r.get("status") == "present")
    absent = sum(1 for r in records if r.get("status") == "absent")
    half_days = sum(1 for r in records if r.get("status") == "half-day")
    on_leave = sum(1 for r in records if r.get("status") == "on-leave")
    
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    ws[f'A{summary_row}'] = "Summary"
    ws[f'A{summary_row}'].font = Font(bold=True, size=11)
    ws[f'A{summary_row}'].fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
    
    ws[f'A{summary_row+1}'] = "Present Days:"
    ws[f'A{summary_row+1}'].font = Font(bold=True)
    ws[f'B{summary_row+1}'] = present
    
    ws[f'D{summary_row+1}'] = "Absent Days:"
    ws[f'D{summary_row+1}'].font = Font(bold=True)
    ws[f'E{summary_row+1}'] = absent
    
    ws[f'A{summary_row+2}'] = "Half Days:"
    ws[f'A{summary_row+2}'].font = Font(bold=True)
    ws[f'B{summary_row+2}'] = half_days
    
    ws[f'D{summary_row+2}'] = "On Leave:"
    ws[f'D{summary_row+2}'].font = Font(bold=True)
    ws[f'E{summary_row+2}'] = on_leave
    
    ws[f'A{summary_row+3}'] = "Total Working Days:"
    ws[f'A{summary_row+3}'].font = Font(bold=True)
    ws[f'B{summary_row+3}'] = len(records)
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Footer
    footer_row = summary_row + 5
    ws[f'A{footer_row}'] = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')} | {company_name}"
    ws[f'A{footer_row}'].font = Font(size=8, color="94a3b8")
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============= API ENDPOINTS =============

@router.get("/download/pdf/{user_id}")
async def download_attendance_pdf(user_id: str, month: Optional[int] = None, year: Optional[int] = None):
    """Download attendance report as PDF for a specific user"""
    if not month:
        month = datetime.now().month
    if not year:
        year = datetime.now().year
    
    records, user = await get_attendance_data(user_id, month, year)
    org_settings = await get_org_settings()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    pdf_buffer = generate_attendance_pdf(records, user, org_settings, month, year)
    
    month_name = calendar.month_name[month]
    filename = f"Attendance_{user.get('name', 'Employee')}_{month_name}_{year}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/download/excel/{user_id}")
async def download_attendance_excel(user_id: str, month: Optional[int] = None, year: Optional[int] = None):
    """Download attendance report as Excel for a specific user"""
    if not month:
        month = datetime.now().month
    if not year:
        year = datetime.now().year
    
    records, user = await get_attendance_data(user_id, month, year)
    org_settings = await get_org_settings()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    excel_buffer = generate_attendance_excel(records, user, org_settings, month, year)
    
    month_name = calendar.month_name[month]
    filename = f"Attendance_{user.get('name', 'Employee')}_{month_name}_{year}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/users")
async def get_all_users_for_attendance():
    """Get list of all users for HR attendance management"""
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users


@router.get("/summary")
async def get_attendance_summary(month: Optional[int] = None, year: Optional[int] = None, department: Optional[str] = None):
    """Get attendance summary for all users (HR view)"""
    if not month:
        month = datetime.now().month
    if not year:
        year = datetime.now().year
    
    all_data = await get_all_users_attendance(month, year, department)
    return {
        "data": all_data,
        "month": month,
        "year": year
    }


@router.get("/download/bulk/excel")
async def download_bulk_attendance_excel(month: Optional[int] = None, year: Optional[int] = None, department: Optional[str] = None):
    """Download attendance report for all employees as Excel"""
    if not month:
        month = datetime.now().month
    if not year:
        year = datetime.now().year
    
    all_data = await get_all_users_attendance(month, year, department)
    org_settings = await get_org_settings()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Employees Attendance"
    
    # Styles
    header_font = Font(bold=True, size=14, color="1e293b")
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    table_header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    # Company Header
    company_name = org_settings.get('name', 'Smarthub Enerzia')
    ws.merge_cells('A1:G1')
    ws['A1'] = company_name
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    month_name = calendar.month_name[month]
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Employee Attendance Summary - {month_name} {year}"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Table Headers
    headers = ['#', 'Employee Name', 'Department', 'Present', 'Absent', 'Half Days', 'On Leave']
    header_row = 5
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for idx, item in enumerate(all_data, 1):
        row = header_row + idx
        user = item.get('user', {})
        summary = item.get('summary', {})
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=user.get('name', 'N/A')).border = thin_border
        ws.cell(row=row, column=3, value=user.get('department', 'N/A')).border = thin_border
        ws.cell(row=row, column=4, value=summary.get('present', 0)).border = thin_border
        ws.cell(row=row, column=5, value=summary.get('absent', 0)).border = thin_border
        ws.cell(row=row, column=6, value=summary.get('half_days', 0)).border = thin_border
        ws.cell(row=row, column=7, value=summary.get('on_leave', 0)).border = thin_border
        
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    
    # Footer
    footer_row = header_row + len(all_data) + 2
    ws[f'A{footer_row}'] = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')} | {company_name}"
    ws[f'A{footer_row}'].font = Font(size=8, color="94a3b8")
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"All_Employees_Attendance_{month_name}_{year}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
